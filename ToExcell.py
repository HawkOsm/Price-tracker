# ToExcell.py
from openpyxl import Workbook, load_workbook
import os

XLSX_PATH = "products.xlsx"
SHEET_NAME = "Products"

def ensure_workbook(path=XLSX_PATH, sheet_name=SHEET_NAME):
    """
    Make sure the workbook and target sheet exist, with a header row.
    """
    if os.path.exists(path):
        wb = load_workbook(path)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["Item", "URL", "Price"])
            wb.save(path)
        else:
            ws = wb[sheet_name]
            # If the sheet exists but is empty, write header
            if ws.max_row == 1 and all((c.value is None for c in ws[1])):
                ws.append(["Item", "URL", "Price"])
                wb.save(path)
        return wb
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["Item", "URL", "Price"])
        wb.save(path)
        return wb

def excel(name, url, *_ignore):
    """
    Append a single product row to the XLSX.
    Kept the same function name so GUI.py doesn't need to change imports.
    Extra args are ignored to stay backward-compatible with old calls.
    """
    ensure_workbook()
    wb = load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]
    ws.append([name, url, None])  # leave Price empty; Tracker.py will fill it
    wb.save(XLSX_PATH)
    wb.close()

def read_all(path=XLSX_PATH, sheet_name=SHEET_NAME):
    """
    Return list of (name, url) for all data rows (skips header).
    """
    if not os.path.exists(path):
        return []
    wb = load_workbook(path)
    ws = wb[sheet_name]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):  # skip header
        name = r[0] if r and len(r) > 0 else ""
        url  = r[1] if r and len(r) > 1 else ""
        if name or url:
            rows.append((name, url))
    wb.close()
    return rows

# Maintain a 'products' symbol so old GUI shutdown code won't crash even if left in.
class _NoOpProducts:
    def close(self):  # xlsxwriter used to require this; openpyxl doesn't
        pass

products = _NoOpProducts()
