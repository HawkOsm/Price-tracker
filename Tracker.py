from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import time

def fetch_price(session, url):
    try:
        resp = session.get(url, timeout=15)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")

        # Try a few common patterns
        cand = (
            soup.select_one("span.price") or
            soup.select_one("span.current-price") or
            soup.select_one("[data-test='price']") or
            soup.select_one(".price .amount") or
            soup.find("meta", itemprop="price")
        )
        if cand is None:
            return None

        if getattr(cand, "get", None) and cand.get("content"):
            return cand["content"].strip()

        text = cand.get_text(strip=True)
        return text or None
    except Exception:
        return None

def update_prices_in_excel(path="./products.xlsx", sheet_name=None, delay_s=0.5):

    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name else wb.active

    with requests.Session() as session:
        # enumerate gives us the row index; each iterated item is a 1-tuple (the B cell value)
        for r_idx, (url,) in enumerate(
            ws.iter_rows(min_row=1, min_col=2, max_col=2, values_only=True),
            start=1
        ):
            if not url:
                continue

            price = fetch_price(session, url)
            ws.cell(row=r_idx, column=3, value=price)   # write into column C

            # Optional: be polite to sites / avoid rate limiting
            if delay_s:
                time.sleep(delay_s)

    wb.save(path)
    wb.close()
    print("Done.")

if __name__ == "__main__":
    update_prices_in_excel("./products.xlsx")
